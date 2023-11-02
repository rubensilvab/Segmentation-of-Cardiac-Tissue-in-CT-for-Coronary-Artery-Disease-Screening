# -*- coding: utf-8 -*-
"""
Created on Wed Nov  2 15:59:51 2022

@author: RubenSilva
"""

# import xlsxwriter module
import xlsxwriter
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import os
import glob
import numpy as np
import shutil
import openpyxl
import cv2
import pydicom

from TestGenerator import *
 
"""Função para fazer predict de um conjunto de imagens"""
def single_dice_coef(y_true, y_pred_bin):
    # shape of y_true and y_pred_bin: (height, width)
    intersection = np.sum(y_true * y_pred_bin)
    if (np.sum(y_true)==0) and (np.sum(y_pred_bin)==0):
        return 1
    return round((2*intersection) / (np.sum(y_true) + np.sum(y_pred_bin)),3)

def predict(model,X,X_original,Width,Length):
  prediction=[]
  for i in range(len(X)):
    pred=model.predict(X[i].reshape(1,Width,Length,3),verbose=0)
    "Reshape para calcular direito MAD e HD"
    pred=np.squeeze(pred)
    woriginal,loriginal=X_original[i].shape[0],X_original[i].shape[1]
    pred=cv2.resize(pred, (woriginal,loriginal))
    pred=(pred>0.5).astype(np.uint8)
    prediction.append(pred)
  return np.array(prediction)

def generators(datagen_image):
    for batch_image in datagen_image:
        yield batch_image

def save_img_results(test_dataframe,path_to,model,Width,Length,change_hu=False):
    patients=np.unique(test_dataframe['Patient'])
    n_patients=len(np.unique(test_dataframe['Patient']))
    
    
    for patient in range(n_patients): 
        
       """Path to go predicts"""
       
       path=os.path.join(path_to,str(patients[patient]))
       isExist = os.path.exists(path)
       #print(path_to_cpy,isExist)
       
       if not isExist:                         
           # Create a new directory because it does not exist 
           os.makedirs(path)
       os.chdir(path)   
       
       input_col = 'Path_image'
       mask_col = 'Path_Mask'
       cols=[input_col,mask_col]
       batch_size=12
       csv_file=test_dataframe.loc[test_dataframe['Patient'].isin([(patients[patient])])]
       data_generator=generators(CustomDataGenerator(csv_file,cols, batch_size=batch_size,input_shape=(Width,Length),change_hu=change_hu))
       s=0
       print("Faltam as imagens previstas para ",n_patients-patient," pacientes, atual:",str(patients[patient]))

       while True:
           try:
               X, Y, X_original = next(data_generator)         
               
               pred_test=predict(model,X,X_original,Width,Length)
               #print(len(X))
               for i in range (len(X)):
                s=s+1 
                fig=plt.figure(figsize=(16,6))
                fig.suptitle('Dice:'+str(round(single_dice_coef(Y[i], np.squeeze(pred_test[i])),3)))
                plt.subplot(1,3,1)
                plt.imshow(np.squeeze(X[i][:,:,1]),cmap='gray')
                plt.title('Original Teste_'+str(s))
                plt.subplot(1,3,2)
                plt.imshow(np.squeeze(Y[i]),cmap='gray')
                plt.title('label Test_'+str(s))
                plt.subplot(1,3,3)
                plt.imshow(np.squeeze(pred_test[i]),cmap='gray')
                plt.title('Predict_'+str(s))
                fig.savefig('Predicts_test_'+str(patients[patient])+"_"+str(s)+'.jpg')
                plt.close('all')
               
        # do something with the batches
           except StopIteration:
        # stop the loop when the generator raises StopIteration
                break      

def excel_results(test_dataframe,model,name,path,Width,Length,change_hu=False):
    """Load do excel modelo"""
    file_excel="X:/Ruben/TESE/New_training_Unet/Models_only_pub_dset/Mix_datasets/excel_modelo.xlsx"
    
    isExist = os.path.exists(path)
    
    if not isExist:                         
        # Create a new directory because it does not exist 
        os.makedirs(path)
       
    
    shutil.copy(file_excel, path)
    old_name=os.path.join(path,"excel_modelo.xlsx")
    new_name=os.path.join(path,str(name)+'.xlsx')
    os.rename(old_name,new_name)
    
    """Overwrite"""
    os.chdir(path)
    book = openpyxl.load_workbook(str(name)+'.xlsx')
    active_sheet = book.active
    
    #workbook = xlsxwriter.Workbook(str(name)+'.xlsx')
     
    # By default worksheet names in the spreadsheet will be
    # Sheet1, Sheet2 etc., but we can also specify a name.
    
    sheet = book.get_sheet_by_name("Results 256")
    #worksheet = workbook.add_worksheet("Results 256")
     
    # Some data we want to write to the worksheet.

    
    # Start from the first cell. Rows and
    # columns are zero indexed.
    row = 1
    col = 1
    
    sheet.cell(row=row, column=col).value = "patient"
    sheet.cell(row=row, column=col+1).value = "Dice_3D"
     
    # Iterate over the data and write it out row by row.
    patients=np.unique(test_dataframe['Patient'])
    
    for patient, i in zip(patients,range(len(patients))):
        row += 1
        
        input_col = 'Path_image'
        mask_col = 'Path_Mask'
        cols=[input_col,mask_col]
        batch_size=12
        csv_file=test_dataframe.loc[test_dataframe['Patient'].isin([(patient)])]
        data_generator=generators(CustomDataGenerator(csv_file,cols, batch_size=batch_size,input_shape=(Width,Length),change_hu=change_hu))
        print("Faltam os resultados do excel para ",len(patients)-i," pacientes, atual: ",str(patient))
       
        X_p=[]
        X_p_original=[]
        Y_p=[]
        
        while True:
            try:
                X, Y, X_original = next(data_generator)         
                X_p.append(np.array(X))
                X_p_original.append(np.array(X_original))
                Y_p.append(np.array(Y))
            
            except StopIteration:
             # stop the loop when the generator raises StopIteration
                     break  
                 
        X_p,X_p_original,Y_p=np.concatenate(X_p, axis=0),np.concatenate(X_p_original,axis=0),np.concatenate(Y_p,axis=0)
        
               
        pred_test=predict(model,X_p,X_p_original,Width,Length)
        
        dice=single_dice_coef(Y_p,np.squeeze(pred_test))
        sheet.cell(row=row, column=col).value = patient
        sheet.cell(row=row, column=col+1).value = dice
       
          
    book.save(str(name)+'.xlsx')

"""Convert to NRRD to view in 3Dslicer"""

def reshape_nrrd(nrrd,n):
  if n==1:  
      nrrd=nrrd[::-1] 
  nrrd=np.transpose(nrrd,(2,1,0))  
 
  return nrrd  

import cc3d

"Funçoes pos processamento"

def fill_3d(labels_out):
    mask_convex=np.zeros(labels_out.shape)
    for i in range(labels_out.shape[0]):
        if( i==0 or i==labels_out.shape[0]-1):
            mask_convex[i,:,:]=labels_out[i,:,:]
        else:
            mask_convex[i,:,:]=np.logical_or(labels_out[i,:,:],np.logical_and(labels_out[i-1,:,:], labels_out[i+1,:,:]))
    return mask_convex.astype(np.uint8)

def connected_components(pred_test):
     # Get a labeling of the k largest objects in the image.
     # The output will be relabeled from 1 to N.
     labels_out, N = cc3d.largest_k(
       pred_test, k=1, 
       connectivity=6, delta=0,
       return_N=True,
     )
    
     labels_out=labels_out.astype(np.uint8)
     
     return labels_out
 
import cv2
import skimage.morphology, skimage.data

def fill_holes(labels_out):
    mask_fill=labels_out.copy()
    filled_img = np.zeros_like(mask_fill)
    for i in range(labels_out.shape[0]):
      # Find contours
        contours, _ = cv2.findContours(mask_fill[i], cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        
        # Draw contours on black image
        contour_img = np.zeros_like(mask_fill[i])
        cv2.drawContours(contour_img, contours, -1, 1, 1)
        
        # Fill enclosed regions
        
        for contour in contours:
            cv2.fillPoly(filled_img[i], [contour], 1)
          
    return filled_img.astype(np.uint8)


from scipy.spatial import ConvexHull

from PIL import Image, ImageDraw

def convex_hull_image(data):
    w,l=data.shape[0],data.shape[1]
    region = np.argwhere(data)
    try:   
        hull = ConvexHull(region)
        verts = [(region[v,0], region[v,1]) for v in hull.vertices]
        img = Image.new('L', data.shape, 0)
        ImageDraw.Draw(img).polygon(verts, outline=1, fill=1)
        mask = np.array(img)
    except:    
        mask=np.zeros((w,l))
    return mask.T

def convex_mask(labels_out):
    mask_convex=np.zeros(labels_out.shape)
    for i in range(labels_out.shape[0]):
            mask_convex[i,:,:]=convex_hull_image(labels_out[i,:,:])
    return mask_convex.astype(np.uint8)


def pos_process(pred):
    connected=connected_components(pred)   
    
    convex_predict=fill_3d(connected)
    fill_2d_predict=fill_3d(connected)
    
    convex_predict=convex_mask(convex_predict)
    fill_2d_predict=fill_holes(fill_2d_predict)
    
   
    return convex_predict,fill_2d_predict


import nrrd
from collections import OrderedDict

def convert_nrrd(test_dataframe,model,path,Width,Length,change_hu=False):
    #Copy header
    if 'ospit' in path:
        print('entrou ceerto')
        PATH_header ='X:/Ruben/TESE/Data/hospital_gaia/EPICHEART_Carolina/selection'
        n=0
    elif 'ardiac' in path:
        
        PATH_header='X:/Ruben/TESE/Data/Dataset_public/RioFatSegm/RioFatSegm/Dicom _ Treino'
        n=1
    else:
        print('entrou errado')
        PATH_header='X:/Ruben/TESE/Data/Dataset_public/Orcya/nrrd_heart' 
        csv_thick=pd.read_csv('X:/Ruben/TESE/New_training_Unet/all_data/OSIC_new/OSIC_3D_test_set_thickness.csv')

        n=1
    
    patients=np.unique(test_dataframe['Patient'])
    
    for i in range(len(patients)):
        
        "CSV contendo os dados dos pacientes"
        
        csv_file=test_dataframe.loc[test_dataframe['Patient'].isin([(patients[i])])]
        
        "Buscar header original"
        if 'ospit' in path:
            PATH_header_nrrd =os.path.join(PATH_header, str(patients[i]), "segm_manual_Carolina.nrrd")
            header = nrrd.read_header(PATH_header_nrrd) 
            
        elif 'ardiac' in path:
            
            PATH_header_nrrd=os.path.join(PATH_header, str(patients[i]))
            files=sorted(glob.glob(PATH_header_nrrd+'/*'))
            # Read the data back from file
            slic=len(csv_file)-1
            data = pydicom.read_file(files[slic])
            pix_spacing= data.get("PixelSpacing")
            pix_spacing[0],pix_spacing[1]=float(pix_spacing[0]),float(pix_spacing[1])
            thick=data.get('SliceThickness')
            origin=data.get('ImagePositionPatient')
            origins=np.array([origin[0],origin[1],origin[2]])
            # print(thick)
            # if thick=='None':
            #     thick=data.get('SliceThickness')
            pix_spacing.append(thick)
            space_directions=np.diag(pix_spacing)
            
            PATH_header_copy ='X:/Ruben/TESE/Data/hospital_gaia/EPICHEART_Carolina/selection' 
            PATH_header_nrrd =os.path.join(PATH_header_copy, str(107780), "segm_manual_Carolina.nrrd")
            header=nrrd.read_header(PATH_header_nrrd)
            header['space directions']=space_directions
            header['space origin']=origins
            
            #print(header)
         
        else:  
            path_dcm="X:/Ruben/TESE/Data/Dataset_public/Orcya/orcic/"
            
            PATH_header_nrrd =os.path.join(PATH_header, str(patients[i]).upper()+"_heart.nrrd")
            header = nrrd.read_header(PATH_header_nrrd) 
            
            #Buscar informaçao dicom
            dicom_files= os.path.join(path_dcm,str(patients[i]))
            files=sorted(glob.glob(dicom_files+'/*'))
            # Read the data back from file
            data = pydicom.read_file(files[0])
            thick_p=csv_thick.loc[csv_thick['Patient'].isin([(patients[i])])]
            thick_p=thick_p['Thickness']
            #thick=data.get('SliceThickness')
            header['space directions'][2][2]=float(thick_p)
            header['space directions']=abs(header['space directions'])
        
        
        
        "Predicts and generator"
        
        input_col = 'Path_image'
        mask_col = 'Path_Mask'
        cols=[input_col,mask_col]
        batch_size=12
        
        data_generator=generators(CustomDataGenerator(csv_file,cols, batch_size=batch_size,input_shape=(Width,Length),change_hu=change_hu))
        
        print("Faltam os dados em nrrd para ",len(patients)-i," pacientes, atual: ",str(patients[i]))
       
        X_p=[]
        X_p_original=[]
        Y_p=[]
        
        while True:
            try:
                X, Y, X_original = next(data_generator)         
                X_p.append(np.array(X))
                X_p_original.append(np.array(X_original))
                Y_p.append(np.array(Y))
            
            except StopIteration:
             # stop the loop when the generator raises StopIteration
                     break  
                 
        X_p,X_p_original,Y_p=np.concatenate(X_p, axis=0),np.concatenate(X_p_original,axis=0),np.concatenate(Y_p,axis=0)
        
        #print(X_p_original.shape,Y_p.shape,np.unique(Y_p))        
        pred_test=predict(model,X_p,X_p_original,Width,Length)
        
        pred_test=np.squeeze(pred_test)
        
        #Connected components and fill slices:
        convex_predict,fill_2d_predict=pos_process(pred_test)  
        convex_predict,fill_2d_predict=convex_predict.astype(pred_test.dtype),fill_2d_predict.astype(pred_test.dtype)
        
        #print(np.unique(PosProcess),PosProcess.dtype)
        
        path_to=os.path.join(path,str(patients[i]))
        isExist = os.path.exists(path_to)
        if not isExist:                         
            # Create a new directory because it does not exist 
            os.makedirs(path_to)
            
        os.chdir(path_to)  
        
        nrrd.write(str(patients[i])+'_'+str(Width)+'.nrrd', reshape_nrrd(X_p_original,n),header=header)
        
        nrrd.write(str(patients[i])+'_'+str(Width)+'_2.5UNet'+'.nrrd', reshape_nrrd(pred_test,n),header=header)
    
        nrrd.write(str(patients[i])+'_'+str(Width)+'_manual'+'.nrrd', reshape_nrrd(Y_p,n),header=header)

        #nrrd.write(str(patients[i])+'_'+str(Width)+'_2.5UNet+Posprss'+'.nrrd', reshape_nrrd(PosProcess,n),header=header)
        
        nrrd.write(str(patients[i])+'_'+str(Width)+'_2.5UNet+pp+conv'+'.nrrd', reshape_nrrd(convex_predict,n),header=header)

        nrrd.write(str(patients[i])+'_'+str(Width)+'_2.5UNet+pp+fill2d'+'.nrrd', reshape_nrrd(fill_2d_predict,n),header=header)
       
 