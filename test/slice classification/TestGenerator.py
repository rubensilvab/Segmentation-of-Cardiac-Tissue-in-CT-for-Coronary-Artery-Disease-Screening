# -*- coding: utf-8 -*-
"""
Created on Mon Apr  3 16:51:40 2023

@author: RubenSilva
"""
# -*- coding: utf-8 -*-
"""
Created on Mon Mar 27 20:36:32 2023

@author: RubenSilva
"""

import os
import numpy as np
import pandas as pd
from tensorflow.keras.utils import Sequence
from tensorflow.keras.preprocessing.image import load_img, img_to_array
import glob
import tensorflow as tf
from tensorflow.keras.preprocessing.image import ImageDataGenerator
import cv2

import xlsxwriter
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import os
import glob
import numpy as np
import shutil
import openpyxl
import cv2
import pydicom
 

class CustomDataGenerator(Sequence):
    
    def __init__(self, csv_file, batch_size=12,input_shape=(256, 256), change_hu=False,shuffle=False, seed=None ):
        
        self.csv_file = csv_file
        self.batch_size = batch_size
        self.input_shape = input_shape
        self.shuffle = shuffle
        self.change_hu=change_hu
        self.list_paths = list(self.csv_file['Path_image'])
        self.list_labels = list(self.csv_file['Label'])
        #print(self.list_labels)
        #print(self.list_paths)
    def __len__(self):
       
        return int(np.ceil(len(self.csv_file) / self.batch_size))

    def __getitem__(self, index):
        # Get the batch indexes
        indexes = range(index * self.batch_size, min((index + 1) * self.batch_size, len(self.list_paths)))
        # Initialize the batch data
        
        batch_x = []
        batch_y = []
        batch_slices=[]
        
        for i, idx in enumerate(indexes):
            # Load the input image and mask from the list
            
            input_path = self.list_paths[idx]
            input_labels=self.list_labels[idx]
            # Load the input image and mask from the CSV file
            
            #Load image
            input_img = cv2.resize(cv2.imread(input_path, flags=cv2.IMREAD_ANYDEPTH), self.input_shape)
            
            w,l=input_img.shape
            
            if self.change_hu:
                input_img=self.reconvert_HU(input_img,50,350)
           
            #Normalize
            input_img=(input_img/65535).astype(np.float32)
               
            batch_x.append(input_img.reshape(w,l,1))
            batch_y.append(input_labels)
            batch_slices.append(input_path)
            
        return np.array(batch_x),np.array(batch_y),batch_slices

    def sort_specific(self,files):
      sorted_files=[]
      for file in files:
             order=file[-7:-3]
             if order[1]=='_':
                 sorted_files.append(file)
      for file in files:
             order=file[-7:-3]
             if order[0]=="_":
                 sorted_files.append(file)  
      for file in files:
             order=file[-8:-3]
             if order[0]=="_":
                 sorted_files.append(file)  
      return sorted_files  
  
    def reconvert_HU(self,array,window_center=50,window_width=350,L=0,W=2000,rescale=True):
       
       img_min = L - W//2 #minimum HU level, escala em HU da imagem original
       img_max =L + W//2 #maximum HU level
       
       reconvert_img=(array/65535)*(img_max - img_min) +  img_min # Reconvertido para HU
       
       
       new_img_min = window_center - window_width//2 #minimum HU level, que pretendemos truncar
       new_img_max =window_center + window_width//2 #maximum HU level, que pretendemos truncar
       
       reconvert_img[reconvert_img<new_img_min] = new_img_min #set img_min for all HU levels less than minimum HU level
       reconvert_img[reconvert_img>new_img_max] = new_img_max #set img_max for all HU levels higher than maximum HU level
       
       if rescale: 
           reconvert_img = (reconvert_img - new_img_min) / (new_img_max - new_img_min)*65535 # Para 16 bit
           
           
       return reconvert_img



 
# """Função para fazer predict de um conjunto de imagens"""
# def single_dice_coef(y_true, y_pred_bin):
#     # shape of y_true and y_pred_bin: (height, width)
#     intersection = np.sum(y_true * y_pred_bin)
#     if (np.sum(y_true)==0) and (np.sum(y_pred_bin)==0):
#         return 1
#     return round((2*intersection) / (np.sum(y_true) + np.sum(y_pred_bin)),3)

# def predict(model,X,X_original,Width,Length):
#   prediction=[]
#   for i in range(len(X)):
#     pred=model.predict(X[i].reshape(1,Width,Length,3),verbose=0)
#     "Reshape para calcular direito MAD e HD"
#     pred=np.squeeze(pred)
#     woriginal,loriginal=X_original[i].shape[0],X_original[i].shape[1]
#     pred=cv2.resize(pred, (woriginal,loriginal))
#     pred=(pred>0.5).astype(np.uint8)
#     prediction.append(pred)
#   return np.array(prediction)

# def generators(datagen_image):
#     for batch_image in datagen_image:
#         yield batch_image

# def save_img_results(test_dataframe,path_to,model,Width,Length,change_hu=False):
#     patients=np.unique(test_dataframe['Patient'])
#     n_patients=len(np.unique(test_dataframe['Patient']))
    
    
#     for patient in range(n_patients): 
        
#         """Path to go predicts"""
       
#         path=os.path.join(path_to,str(patients[patient]))
#         isExist = os.path.exists(path)
#         #print(path_to_cpy,isExist)
       
#         if not isExist:                         
#             # Create a new directory because it does not exist 
#             os.makedirs(path)
#         os.chdir(path)   
       
#         input_col = 'Path_image'
#         mask_col = 'Path_Mask'
#         cols=[input_col,mask_col]
#         batch_size=12
#         csv_file=test_dataframe.loc[test_dataframe['Patient'].isin([(patients[patient])])]
#         print(csv_file)
#         data_generator=generators(CustomDataGenerator(csv_file,cols, batch_size=batch_size,input_shape=(Width,Length),change_hu=change_hu))
#         s=0
#         print("Faltam as imagens previstas para ",n_patients-patient," pacientes, atual:",str(patients[patient]))

#         while True:
#             try:
#                 X, Y, X_original = next(data_generator)         
               
#                 pred_test=predict(model,X,X_original,Width,Length)
#                 #print(len(X))
#                 for i in range (len(X)):
#                     s=s+1 
#                     fig=plt.figure(figsize=(16,6))
#                     fig.suptitle('Dice:'+str(round(single_dice_coef(Y[i], np.squeeze(pred_test[i])),3)))
#                     plt.subplot(1,3,1)
#                     plt.imshow(np.squeeze(X[i][:,:,1]),cmap='gray')
#                     plt.title('Original Teste_'+str(s))
#                     plt.subplot(1,3,2)
#                     plt.imshow(np.squeeze(Y[i]),cmap='gray')
#                     plt.title('label Test_'+str(s))
#                     plt.subplot(1,3,3)
#                     plt.imshow(np.squeeze(pred_test[i]),cmap='gray')
#                     plt.title('Predict_'+str(s))
#                     fig.savefig('Predicts_test_'+str(patients[patient])+"_"+str(s)+'.jpg')
#                     plt.close('all')
               
#         # do something with the batches
#             except StopIteration:
#         # stop the loop when the generator raises StopIteration
#                 break      

# def excel_results(test_dataframe,model,name,path,Width,Length,change_hu=False):
#     """Load do excel modelo"""
#     file_excel="X:/Ruben/TESE/New_training_Unet/Models_only_pub_dset/Mix_datasets/excel_modelo.xlsx"
#     shutil.copy(file_excel, path)
#     old_name=os.path.join(path,"excel_modelo.xlsx")
#     new_name=os.path.join(path,str(name)+'.xlsx')
#     os.rename(old_name,new_name)
    
#     """Overwrite"""
#     os.chdir(path)
#     book = openpyxl.load_workbook(str(name)+'.xlsx')
#     active_sheet = book.active
    
#     #workbook = xlsxwriter.Workbook(str(name)+'.xlsx')
     
#     # By default worksheet names in the spreadsheet will be
#     # Sheet1, Sheet2 etc., but we can also specify a name.
    
#     sheet = book.get_sheet_by_name("Results 256")
#     #worksheet = workbook.add_worksheet("Results 256")
     
#     # Some data we want to write to the worksheet.

    
#     # Start from the first cell. Rows and
#     # columns are zero indexed.
#     row = 1
#     col = 1
    
#     sheet.cell(row=row, column=col).value = "patient"
#     sheet.cell(row=row, column=col+1).value = "Dice_3D"
     
#     # Iterate over the data and write it out row by row.
#     patients=np.unique(test_dataframe['Patient'])
    
#     for patient, i in zip(patients,range(len(patients))):
#         row += 1
        
#         input_col = 'Path_image'
#         mask_col = 'Path_Mask'
#         cols=[input_col,mask_col]
#         batch_size=12
#         csv_file=test_dataframe.loc[test_dataframe['Patient'].isin([str(patient)])]
#         data_generator=generators(CustomDataGenerator(csv_file,cols, batch_size=batch_size,input_shape=(Width,Length),change_hu=change_hu))
#         print("Faltam os resultados do excel para ",len(patients)-i," pacientes, atual: ",str(patient))
       
#         X_p=[]
#         X_p_original=[]
#         Y_p=[]
        
#         while True:
#             try:
#                 X, Y, X_original = next(data_generator)         
#                 X_p.append(np.array(X))
#                 X_p_original.append(np.array(X_original))
#                 Y_p.append(np.array(Y))
            
#             except StopIteration:
#               # stop the loop when the generator raises StopIteration
#                       break  
                 
#         X_p,X_p_original,Y_p=np.concatenate(X_p, axis=0),np.concatenate(X_p_original,axis=0),np.concatenate(Y_p,axis=0)
        
               
#         pred_test=predict(model,X_p,X_p_original,Width,Length)
        
#         dice=single_dice_coef(Y_p,np.squeeze(pred_test))
#         sheet.cell(row=row, column=col).value = patient
#         sheet.cell(row=row, column=col+1).value = dice
       
          
#     book.save(str(name)+'.xlsx')

# """Convert to NRRD to view in 3Dslicer"""

# def reshape_nrrd(nrrd,n):
#   if n==1:  
#       nrrd=nrrd[::-1] 
#   nrrd=np.transpose(nrrd,(2,1,0))  
 
#   return nrrd  

# import nrrd
# from collections import OrderedDict

# def convert_nrrd(test_dataframe,model,path,Width,Length,change_hu=False):
#     #Copy header
#     if 'ospit' in path:
#         print('entrou ceerto')
#         PATH_header ='X:/Ruben/TESE/Data/hospital_gaia/EPICHEART_Carolina/selection'
#         n=0
#     elif 'ardiac' in path:
        
#         PATH_header='X:/Ruben/TESE/Data/Dataset_public/RioFatSegm/RioFatSegm/Dicom _ Treino'
#         n=1
#     else:
#         print('entrou errado')
#         PATH_header='X:/Ruben/TESE/Data/Dataset_public/Orcya/nrrd_heart' 
#         n=1
    
#     patients=np.unique(test_dataframe['Patient'])
    
#     for i in range(len(patients)):
        
#         "CSV contendo os dados dos pacientes"
        
#         csv_file=test_dataframe.loc[test_dataframe['Patient'].isin([str(patients[i])])]
        
#         "Buscar header original"
#         if 'ospit' in path:
#             PATH_header_nrrd =os.path.join(PATH_header, str(patients[i]), "segm_manual_Carolina.nrrd")
#             header = nrrd.read_header(PATH_header_nrrd) 
            
#         elif 'ardiac' in path:
            
#             PATH_header_nrrd=os.path.join(PATH_header, str(patients[i]))
#             files=sorted(glob.glob(PATH_header_nrrd+'/*'))
#             # Read the data back from file
#             slic=len(csv_file)-1
#             data = pydicom.read_file(files[slic])
#             pix_spacing= data.get("PixelSpacing")
#             pix_spacing[0],pix_spacing[1]=float(pix_spacing[0]),float(pix_spacing[1])
#             thick=data.get('SliceThickness')
#             origin=data.get('ImagePositionPatient')
#             origins=np.array([origin[0],origin[1],origin[2]])
#             # print(thick)
#             # if thick=='None':
#             #     thick=data.get('SliceThickness')
#             pix_spacing.append(thick)
#             space_directions=np.diag(pix_spacing)
            
#             PATH_header_copy ='X:/Ruben/TESE/Data/hospital_gaia/EPICHEART_Carolina/selection' 
#             PATH_header_nrrd =os.path.join(PATH_header_copy, str(107780), "segm_manual_Carolina.nrrd")
#             header=nrrd.read_header(PATH_header_nrrd)
#             header['space directions']=space_directions
#             header['space origin']=origins
            
#             #print(header)
         
#         else:  
#             path_dcm="X:/Ruben/TESE/Data/Dataset_public/Orcya/orcic/"
            
#             PATH_header_nrrd =os.path.join(PATH_header, str(patients[i]).upper()+"_heart.nrrd")
#             header = nrrd.read_header(PATH_header_nrrd) 
            
#             #Buscar informaçao dicom
#             dicom_files= os.path.join(path_dcm,str(patients[i]))
#             files=sorted(glob.glob(dicom_files+'/*'))
#             # Read the data back from file
#             data = pydicom.read_file(files[0])
#             thick=data.get('SliceThickness')
#             header['space directions'][2][2]=float(thick)
#             header['space directions']=abs(header['space directions'])
        
        
        
#         "Predicts and generator"
        
#         input_col = 'Path_image'
#         mask_col = 'Path_Mask'
#         cols=[input_col,mask_col]
#         batch_size=12
        
#         data_generator=generators(CustomDataGenerator(csv_file,cols, batch_size=batch_size,input_shape=(Width,Length),change_hu=change_hu))
        
#         print("Faltam os dados em nrrd para ",len(patients)-i," pacientes, atual: ",str(patients[i]))
       
#         X_p=[]
#         X_p_original=[]
#         Y_p=[]
        
#         while True:
#             try:
#                 X, Y, X_original = next(data_generator)         
#                 X_p.append(np.array(X))
#                 X_p_original.append(np.array(X_original))
#                 Y_p.append(np.array(Y))
            
#             except StopIteration:
#               # stop the loop when the generator raises StopIteration
#                       break  
                 
#         X_p,X_p_original,Y_p=np.concatenate(X_p, axis=0),np.concatenate(X_p_original,axis=0),np.concatenate(Y_p,axis=0)
        
#         #print(X_p_original.shape,Y_p.shape,np.unique(Y_p))        
#         pred_test=predict(model,X_p,X_p_original,Width,Length)
        
#         pred_test=np.squeeze(pred_test)
        
#         path_to=os.path.join(path,str(patients[i]))
#         isExist = os.path.exists(path_to)
#         if not isExist:                         
#             # Create a new directory because it does not exist 
#             os.makedirs(path_to)
            
#         os.chdir(path_to)  
        
#         nrrd.write(str(patients[i])+'_'+str(Width)+'.nrrd', reshape_nrrd(X_p_original,n),header=header)
        
#         nrrd.write(str(patients[i])+'_'+str(Width)+'_2.5UNet'+'.nrrd', reshape_nrrd(pred_test,n),header=header)
    
#         nrrd.write(str(patients[i])+'_'+str(Width)+'_manual'+'.nrrd', reshape_nrrd(Y_p,n),header=header)

   
    
 
# """See results"""

# Width,Length=256,256
# from keras.models import Model
# from keras.models import load_model

# #cfat_peri_df = pd.read_csv('data_only_pericardium/Cardiac_fat_new/Cardiac_fat_new_folds_5.csv')
# cfat_all_df = pd.read_csv('X:/Ruben/TESE/New_training_Unet/all_data/Cardiac_fat_new/Cardiac_fat_new_folds_5.csv')
# #osic_peri_df = pd.read_csv('X:/Ruben/TESE/New_training_Unet/data_only_pericardium/OSIC_new/OSIC_new_folds_5.csv')
# osic_all_df = pd.read_csv('X:/Ruben/TESE/New_training_Unet/all_data/OSIC_new/OSIC_new_folds_5.csv')

# test_hospit_df= pd.read_csv('X:/Ruben/TESE/Data/hospital_gaia/all_data_carolina/all_data_carolina_hospital_1.csv')



# """Load the model"""
# path_model="X:/Ruben/TESE/New_training_Unet/Models_only_pub_dset/Mix_datasets/All_data/models/2.5D_Unet/Dice_loss/L0_W2000_tif/Sun Apr  2 20_57_04 2023"
# os.chdir(path_model)

# model = load_model('Loss_Dice_loss__epochs_4000_batch_size_12_wl256Lr_decreasing_0.0001fold_train_0_1_2_time_Sun_Apr_2_20_57_04_2023.h5',compile=False)


# path_to="X:/Ruben/TESE/New_training_Unet/Models_only_pub_dset/Mix_datasets/All_data/predict/2.5D_Unet/Dice_loss/osic_tif_teste/L0_W2000_tif" 

# """See results"""
# save_img_results(test_hospit_df,path_to,model,Width,Length) 