# -*- coding: utf-8 -*-
"""
Created on Wed Nov  2 15:59:51 2022

@author: RubenSilva
"""

# import xlsxwriter module
import xlsxwriter
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import os
import glob
import numpy as np
import shutil
import openpyxl
import cv2
import pydicom
import pandas as pd

from TestGenerator import *
  
from tensorflow.keras.preprocessing.image import ImageDataGenerator



def generators(datagen_image):
    for batch in datagen_image:
        
        yield batch


"""Função para fazer predict de um conjunto de imagens"""

def predict(model,X,Width,Length):
   
  # for i in range(len(X)):
  #   pred=model.predict(X[i].reshape(1,Width,Length,1),verbose=0)
  #   "Reshape para calcular direito MAD e HD"
    pred=model.predict(X,verbose=0)
    "Retirar para fazer thresold tuning"
    
    pred=(pred>0.517566).astype(np.uint8)
    
    return pred

def get_excel_row(path,patient,name,slic,label,pred):
    
    isExist = os.path.exists(path)
    if not isExist:                         
        # Create a new directory because it does not exist 
        os.makedirs(path)
        
    os.chdir(path) 
    
        
    filename = str(name)+'.xlsx'
    
    # Check if the file already exists
    if not os.path.isfile(filename):
        # Create a new workbook object
        book = openpyxl.Workbook()
    
        # Select the worksheet to add data to
        sheet = book.active
    
        # Add a header row to the worksheet
        sheet.append(['Patient', 'Slice', 'Label', 'Prediction'])
    else:
        # Open an existing workbook
        book = openpyxl.load_workbook(filename)
    
        # Select the worksheet to add data to
        sheet = book.active
    
    # Append a new row of data to the worksheet
    sheet.append([patient, slic, label, pred])
    #print([patient, slic, label, pred])
    # Save the workbook to a file
    book.save(filename)


def excel_results(test_dataframe,model,name,path,Width,Length,change_hu=False):
    
    # Iterate over the data and write it out row by row.
    patients=np.unique(test_dataframe['Patient'])
    #test_dataframe ['Label'] = test_dataframe['Label'].astype(str)

    for patient, i in zip(patients,range(len(patients))):
        
    
        batch_size=12
        csv_file=test_dataframe.loc[test_dataframe['Patient'].isin([(patient)])]
        data_generator=generators(CustomDataGenerator(csv_file, batch_size,input_shape=(Width,Length),change_hu=change_hu))
        print("Faltam os resultados do excel para ",len(patients)-i," pacientes, atual: ",str(patient))
       
    
        while True:
            try:
                batch_x,batch_y,paths  = next(data_generator)
                slices_name=paths
                pred_test=predict(model,batch_x,Width,Length)
                #print(pred_test)
                #print(batch_y.shape)
                
                for sli in range(batch_y.shape[0]):
                    
                    get_excel_row(path,patient,name,slices_name[sli],batch_y[sli],pred_test[sli][0])

            
            except StopIteration:
             # stop the loop when the generator raises StopIteration
                     break  


import seaborn as sns
from sklearn.metrics import confusion_matrix,classification_report
from openpyxl.styles import PatternFill

def get_excel_row_report(patient,name,df):
    
    filename = str(name)+'.xlsx'
    
    # Check if the file already exists
    if not os.path.isfile(filename):
        # Create a new workbook object
        book = openpyxl.Workbook()
    
        # Select the worksheet to add data to
        sheet = book.active
        # Merge cells
        sheet.merge_cells("B1:E1")
        sheet.merge_cells("F1:I1")
        # Set the value in the merged cells
        sheet["B1"] = "no pericardium"
        sheet["F1"] = "pericardium"
        # Add a header row to the worksheet
        sheet.append(['Patient','support(No P)', 'precision', 'recall', 'f1-score','support(Pericardium)','precision', 'recall', 'f1-score','Accuracy'])
    
    
    else:
        # Open an existing workbook
        book = openpyxl.load_workbook(filename)
    
        # Select the worksheet to add data to
        sheet = book.active
    
    # Append a new row of data to the worksheet
    sheet.append([patient,df['support'][0],df['precision'][0],df['recall'][0],df['f1-score'][0],df['support'][1],df['precision'][1],df['recall'][1],df['f1-score'][1],df['f1-score'][2]])
    #print([patient, slic, label, pred])
    # Save the workbook to a file
    book.save(filename)

def get_ConfusionMatrix(y_true,y_pred,class_names,patient):
    

    # Calculate the confusion matrix
    cm = confusion_matrix(y_true, y_pred)

    # Create a heatmap of the confusion matrix
    plt.figure(figsize=(16, 12),dpi=300)
    sns.heatmap(cm, annot=True, fmt="d", cmap="Blues",annot_kws={"fontsize": 14})
    
    # Add labels, title, and axis ticks
    plt.xlabel("Predicted labels",fontsize=(20),fontweight= "bold")
    plt.ylabel("True labels",fontsize=(20),fontweight= "bold")
    plt.title("Confusion Matrix",fontsize=25,fontweight= "bold")
    plt.xticks(ticks=np.arange(len(class_names))+0.5, labels=class_names,ha='center',fontsize=(15))
    plt.yticks(ticks=np.arange(len(class_names))+0.5, labels=class_names,va='center',fontsize=(15))

  
    # Save the figure as an image
    plt.savefig("CM_"+str(patient)+".png")    

def get_classification_metrics(path,name):
    data=pd.read_excel(os.path.join(path,name+'.xlsx'))
    
    # Iterate over the data and write it out row by row.
    patients=np.unique(data['Patient'])
    #test_dataframe ['Label'] = test_dataframe['Label'].astype(str)
    name_excel='Metrics_Report'
    for patient, i in zip(patients,range(len(patients))):
        
        csv_patient=data.loc[data['Patient'].isin([(patient)])]
        
        print("Faltam os relatórios para ",len(patients)-i," pacientes, atual: ",str(patient))

        y_true=csv_patient['Label']
        y_pred=csv_patient['Prediction']
        # Substitute label values
        class_names = ["No pericardium", "pericardium"]
        
        y_true_names = [class_names[label] for label in y_true]
        y_pred_names = [class_names[label] for label in y_pred]
        
        path_to=os.path.join(path,'Reports')
        
        isExist = os.path.exists(path_to)
        if not isExist:                         
            # Create a new directory because it does not exist 
            os.makedirs(path_to)
            
        os.chdir(path_to) 
        get_ConfusionMatrix(y_true_names,y_pred_names,class_names,patient)
        
        report = classification_report(y_true_names, y_pred_names,output_dict=True)
        # Convert the report dictionary to a DataFrame and transpose it
        report_patient = pd.DataFrame(report).transpose()
        
        
        get_excel_row_report(patient,name_excel,report_patient)
    
    print("Calculando resultados globais...")

    y_true=data['Label']
    y_pred=data['Prediction']
    y_true_names = [class_names[label] for label in y_true]
    y_pred_names = [class_names[label] for label in y_pred]   
    
    patient='All'
    get_ConfusionMatrix(y_true_names,y_pred_names,class_names,patient)
    report = classification_report(y_true_names, y_pred_names,output_dict=True)
    # Convert the report dictionary to a DataFrame and transpose it
    report_patient = pd.DataFrame(report).transpose()
    get_excel_row_report(patient,name_excel,report_patient) 

def fisrt_last_indice(csv):
    #print(csv)
    y_true=csv['Label']
    y_pred=csv['Prediction']
    #First and last indices
    indices=[]
    for i in range(len(csv)):
        #print(i)
        if y_pred.iloc[i]==1:
            #print(y_true[i])
            indices.append(i)
    
    first_idx,last_idx=indices[0],indices[-1]
    #print(first_idx,last_idx)
    
    return first_idx,last_idx
    

def csv_segmentation(df,path,csv_name):
    data=pd.read_excel(os.path.join(path,csv_name+'.xlsx'))
    
    # Iterate over the data and write it out row by row.
    patients=np.unique(data['Patient'])
    #test_dataframe ['Label'] = test_dataframe['Label'].astype(str)
    name_excel=csv_name+'_to_segm'+'.csv'
    csv_to_segm=pd.DataFrame()
    for patient, i in zip(patients,range(len(patients))):
        
        csv_patient=data.loc[data['Patient'].isin([(patient)])]
        csv_patient_ori=df.loc[df['Patient'].isin([(patient)])]
        first_idx,last_idx=fisrt_last_indice(csv_patient)
        
        csv_to_seg_pat=csv_patient_ori.iloc[first_idx:last_idx]
        print("Faltam os csv para segmentar de ",len(patients)-i," pacientes, atual: ",str(patient))
        csv_to_segm=pd.concat([csv_to_segm,csv_to_seg_pat])
       # print(csv_to_segm)
                
    
    csv_to_segm.to_csv(os.path.join(path,name_excel), index=False)
    return  csv_to_segm  
        
# cfat_all_df = pd.read_csv('X:/Ruben/TESE/New_training_Unet/all_data/Cardiac_fat_new/Cardiac_fat_new_folds_5.csv')
# osic_all_df = pd.read_csv('X:/Ruben/TESE/New_training_Unet/all_data/OSIC_new/OSIC_3D_test_set.csv')

# test_cfat_df=cfat_all_df.loc[cfat_all_df['Fold'].isin([4])]
# test_osic_df=osic_all_df.loc[osic_all_df['Fold'].isin([4])]

# path="X:/Ruben/TESE/New_training_Unet/Models_only_pub_dset/Mix_datasets/All_data/classification_models/predict/2D_CNN/BCE/OSIC_tif/L0_W2000_tif_calc_augm"    
# name="results_dice_OSIC_tifL0_W2000_tif_calc_augm.xlsx"

# csv_to_segm=csv_segmentation(test_osic_df,path,name)