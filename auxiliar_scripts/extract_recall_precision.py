# -*- coding: utf-8 -*-
"""
Created on Wed Jul 19 17:45:42 2023

@author: RubenSilva
"""

# import xlsxwriter module
import xlsxwriter
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import os
import glob
import numpy as np
import shutil
import openpyxl
import cv2
import pydicom
import pandas as pd
import seaborn as sns
from sklearn.metrics import confusion_matrix,classification_report
from openpyxl.styles import PatternFill
import nrrd

def reshape_nrrd_to_arr(nrrd):
  
  nrrd=np.array(nrrd)
  nrrd=np.transpose(nrrd,(2,1,0))
  nrrd=nrrd[::-1]  
   
  return nrrd  


def get_excel_row_report(patient,name,df):
    print('Getting excel row')
    filename = str(name)+'.xlsx'
    
    # Check if the file already exists
    if not os.path.isfile(filename):
        # Create a new workbook object
        book = openpyxl.Workbook()
    
        # Select the worksheet to add data to
        sheet = book.active
        # Merge cells
        sheet.merge_cells("B1:E1")
        sheet.merge_cells("F1:I1")
        # Set the value in the merged cells
        sheet["B1"] = "No EAT"
        sheet["F1"] = "EAT"
        # Add a header row to the worksheet
        sheet.append(['Patient','support(No EAT)', 'precision', 'recall', 'f1-score','support(EAT)','precision', 'recall', 'f1-score','Accuracy'])
    
    
    else:
        # Open an existing workbook
        book = openpyxl.load_workbook(filename)
    
        # Select the worksheet to add data to
        sheet = book.active
    
    # Append a new row of data to the worksheet
    sheet.append([patient,df['support'][0],df['precision'][0],df['recall'][0],df['f1-score'][0],df['support'][1],df['precision'][1],df['recall'][1],df['f1-score'][1],df['f1-score'][2]])
    #print([patient, slic, label, pred])
    # Save the workbook to a file
    book.save(filename)

def get_ConfusionMatrix(y_true,y_pred,class_names,patient):
    

    # Calculate the confusion matrix
    print('Getting confusion matrix')
    cm = confusion_matrix(y_true, y_pred)

    # Create a heatmap of the confusion matrix
    plt.figure(figsize=(16, 12),dpi=300)
    sns.heatmap(cm, annot=True, fmt="d", cmap="Blues",annot_kws={"fontsize": 14})
    
    # Add labels, title, and axis ticks
    plt.xlabel("Predicted labels",fontsize=(20),fontweight= "bold")
    plt.ylabel("True labels",fontsize=(20),fontweight= "bold")
    plt.title("Confusion Matrix",fontsize=25,fontweight= "bold")
    plt.xticks(ticks=np.arange(len(class_names))+0.5, labels=class_names,ha='center',fontsize=(15))
    plt.yticks(ticks=np.arange(len(class_names))+0.5, labels=class_names,va='center',fontsize=(15))

  
    # Save the figure as an image
    plt.savefig("CM_"+str(patient)+".png")    

def get_classification_metrics(path_nrrd):
    
    path_reader_2="C:/Users/RubenSilva/Desktop/segmentation_inter_intra/selection/inverted/EAT_segm_nHU/NRRD"
    # Iterate over the data and write it out row by row.
    patients=[patient for patient in os.listdir(path_reader_2) if os.path.isdir(os.path.join(path_reader_2, patient)) ]
    #test_dataframe ['Label'] = test_dataframe['Label'].astype(str)
    name_excel='Metrics_Report'
    
    for patient, i in zip(patients,range(len(patients))):
        
        print('Faltam',len(patients)-i,' pacientes. Atual:',patient)
        #path_nrrd_patient=os.path.join(path_nrrd,patient)
        path_nrrd_reader_2=os.path.join(path_reader_2,patient)
        #buscar nrrds
        #files_nrrd=glob.glob(os.path.join(path_nrrd_patient,'*'))
        files_nrrd_reader2=glob.glob(os.path.join(path_nrrd_reader_2,'*'))

        for file in files_nrrd_reader2:
            
            ##if file[-6]=='v':
               # Convex_mask, header = nrrd.read(file)
                #nrrd_pred=reshape_nrrd_to_arr(Convex_mask)
            
            if 'manual' in file:   
                manual, header = nrrd.read(file)
                manual=reshape_nrrd_to_arr(manual)
            
            # if 'carol' in file:   
            #     carol, header = nrrd.read(file)
            #     nrrd_reader_1=reshape_nrrd_to_arr(carol)
                
            if 'fabio' in file:   
                fabio, header = nrrd.read(file)
                nrrd_reader_1=reshape_nrrd_to_arr(fabio)    
        
        
        
        print("Faltam os relatórios para ",len(patients)-i," pacientes, atual: ",str(patient))

        y_true=manual.flatten()
        y_pred=nrrd_reader_1.flatten()
        print(np.unique(nrrd_reader_1))
        # Substitute label values
        class_names = ["EAT", "No EAT"]
        #print(np.sum(y_true))
        #print(np.sum(y_pred))
        y_true_names = [class_names[label] for label in y_true]
        y_pred_names = [class_names[label] for label in y_pred]
        
        path_to=os.path.join(path_reader_2,'Reports_inter')
        
        isExist = os.path.exists(path_to)
        if not isExist:                         
            # Create a new directory because it does not exist 
            os.makedirs(path_to)
            
        os.chdir(path_to) 
        get_ConfusionMatrix(y_true_names,y_pred_names,class_names,patient)
        
        print('Report calculation')
        report = classification_report(y_true_names, y_pred_names,output_dict=True)
        # Convert the report dictionary to a DataFrame and transpose it
        report_patient = pd.DataFrame(report).transpose()
        
        
        get_excel_row_report(patient,name_excel,report_patient)

path_nrrd="X:/Ruben/TESE/New_training_Unet/Models_only_pub_dset/Mix_datasets/All_data/predict/2.5D_Unet/Dice_loss/Hospital_tif/L0_W2000_calc_augm_tif/EAT_segm_nHU/NRRD"    
get_classification_metrics(path_nrrd)  